import openpyxl
import os
import shutil
import pandas as pd
from datetime import datetime
from database import get_food

TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), 'data', 'Mastersheet_TAJ_CAL27.xlsx')
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), 'data', 'output')

if not os.path.exists(OUTPUT_DIR):
    os.makedirs(OUTPUT_DIR)

def process_bulk_upload_excel(file_path):
    """
    Reads an uploaded Excel file.
    Expected Columns: "Food Name", "Calories", "Allergens"
    Returns: list of dicts [{'name': '...', 'calories': ..., 'allergens': [...]}]
    """
    try:
        df = pd.read_excel(file_path)
    except Exception as e:
        print(f"Error reading Excel: {e}")
        return []

    # clean column names (strip spaces, lower case for matching)
    df.columns = [c.strip().lower() for c in df.columns]
    
    # Map expected columns
    # We expect "food name", "calories", "allergens"
    # Let's map somewhat flexibly
    col_name = next((c for c in df.columns if 'name' in c), None)
    col_cal = next((c for c in df.columns if 'calor' in c), None)
    col_alg = next((c for c in df.columns if 'allergy' in c or 'allergen' in c), None)

    if not col_name or not col_cal:
        print("Required columns (Food Name, Calories) not found")
        return []

    items = []
    for _, row in df.iterrows():
        name = str(row[col_name]).strip()
        if not name or name.lower() == 'nan':
            continue
            
        calories = row[col_cal]
        # Handle NaN calories
        if pd.isna(calories):
            calories = 0
            
        allergens_str = str(row[col_alg]) if col_alg and not pd.isna(row[col_alg]) else ""
        # Split allergens by comma
        allergens = [a.strip() for a in allergens_str.split(',') if a.strip()]
        
        items.append({
            'name': name.upper(), # Enforce Uppercase
            'calories': int(calories),
            'allergens': allergens
        })
        
    return items

def generate_excel(food_names, custom_data=None):
    """
    Generates an Excel file filled with food data.
    food_names: List of strings (food names).
    custom_data: Optional dictionary {'NAME': {'calories': 123, 'allergens': '...'}} to bypass DB.
    Returns: Path to the generated file, and a list of missing foods.
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = os.path.join(OUTPUT_DIR, f"Buffet_Tags_{timestamp}.xlsx")
    
    # Copy template
    shutil.copy(TEMPLATE_PATH, output_file)
    
    wb = openpyxl.load_workbook(output_file)
    ws = wb.active

    # Prepare data to fill
    missing_foods = []
    
    # Limit to first 50 items
    items_to_process = food_names[:50]
    
    # Start row (assuming header is row 1)
    start_row = 2
    max_row = 51 # Header + 50 items
    
    for i, name in enumerate(items_to_process):
        current_row = start_row + i
        clean_name = name.strip().upper()
        if not clean_name:
            continue
            
        if custom_data and clean_name in custom_data:
            food_data = custom_data[clean_name]
        else:
            food_data = get_food(clean_name)
        
        # Column Mappings (1-based index)
        # Food Name: D (4)
        # Calories: W (23)
        # Allergens: X (24) to AK (37)
        
        # Update Name
        ws.cell(row=current_row, column=4, value=clean_name)
        
        # Clear Calories and Allergens for this row strictly (do NOT touch red cols H, I)
        ws.cell(row=current_row, column=23, value="") # Clear W (Calories)
        for col in range(24, 38): # X to AK
             ws.cell(row=current_row, column=col, value="")
        
        if food_data:
            # Fill Calories
            ws.cell(row=current_row, column=23, value=food_data['calories'])
            
            # Allergen Mapping
            # Map DB allergen strings to Column Indices
            allergen_map = {
                'Crustaceans': 24, # X
                'Molluscs': 25,    # Y
                'Fish': 26,        # Z
                'Soy': 27,         # AA - Renamed from Soya
                'Gluten': 28,      # AB
                'Mustard': 29,     # AC
                'Sesame': 30,      # AD
                'Celery': 31,      # AE
                'Eggs': 32,        # AF
                'Milk': 33,        # AG
                'Peanuts': 34,     # AH
                'Nuts': 35,        # AI
                'Sulphite': 36,    # AJ
                'Lupin': 37        # AK
            }
            
            # food_data['allergens'] might be a list (from custom_data) or string (from DB)
            raw_allergens = food_data['allergens']
            if isinstance(raw_allergens, list):
                db_allergens = [str(a).strip() for a in raw_allergens]
            elif isinstance(raw_allergens, str) and raw_allergens:
                db_allergens = [a.strip() for a in raw_allergens.split(',')]
            else:
                db_allergens = []
            
            for allergen in db_allergens:
                # We need to match somewhat loosely or exactly?
                # DB has "Fish", "Sesame seeds" etc from checkbox.
                # Let's try exact match from map keys
                if allergen in allergen_map:
                    col_idx = allergen_map[allergen]
                    ws.cell(row=current_row, column=col_idx, value="yes")
                else:
                    # Fallback check?
                    pass

        else:
            missing_foods.append(clean_name)
    
    # Force delete rows logic DISABLED by user request (2026-02-17)
    # The user reported 999 rows being processed. We must clean up aggressively.
    
    # last_filled_row = start_row + len(items_to_process) - 1
    # start_delete = last_filled_row + 1
    # target_end_row = 1000 # Cover user's reported 999 range + 1
    
    # # Calculate amount to delete.
    # # Even if openpyxl thinks max_row is small, we force delete up to 1000.
    # amount_to_delete = target_end_row - start_delete + 1
    
    # if amount_to_delete > 0:
    #     ws.delete_rows(start_delete, amount_to_delete)

    wb.save(output_file)
    return output_file, missing_foods

def extract_names_from_excel(file_path):
    """
    Extracts values from column D (rows 2 to 60) from the first sheet.
    Ignores empty values.
    Returns: list of strings (food names).
    """
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        
        extracted_names = []
        # Rows 2 to 60 (inclusive), Column 4 (D)
        for row in range(2, 61):
            cell_value = ws.cell(row=row, column=4).value
            if cell_value:
                name = str(cell_value).strip()
                if name and name.lower() != 'nan':
                    extracted_names.append(name)
                    
        return extracted_names
    except Exception as e:
        print(f"Extraction Error: {e}")
        return []
